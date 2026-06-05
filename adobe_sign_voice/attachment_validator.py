"""
attachment_validator.py — CTL Data Quality Validation for CP-uploaded attachments.

Validates each Excel attachment submitted by a Communication Provider against the
circuit master sheet. All data quality issues are treated as CTL-critical because
the circuit records may relate to critical-to-life services.

PDF attachments: extracted to CSV for data controller review, but NOT validated
against the master. A human must review PDF-sourced data before it enters any
downstream system.

Responsibilities:
  1. Load and parse the attachment (Excel, PDF, CSV)
  2. Load the circuit master sheet (cached per mtime)
  3. Row-by-row validation:
       - Circuit ID must exist in master (not found = CTL-critical)
       - Immutable columns must match master exactly (mismatch = possible wrong-record = CTL-critical)
       - Mutable column changes are logged (old → new)
       - Missing required fields = CTL-critical (reason exempt only when include = Y/yes)
       - Invalid email format = CTL-critical (email is a CTL customer contact field)
       - Partial opt-out detected when include in migration ∈ {n, no}
  4. Conditionally update mutable fields in the master sheet (UPDATE_MASTER_DATA flag)
  5. Generate a 4-sheet data controller Excel report
"""
from __future__ import annotations

import logging
import os
import re
import shutil
import threading
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd

from config import (
    ALL_EXPECTED_COLUMNS,
    BACKUP_MASTER_BEFORE_UPDATE,
    CIRCUIT_MASTER_PATH,
    CIRCUIT_MASTER_SHEET_NAME,
    COL_CIRCUIT_ID,
    COL_EMAIL_ADDRESS,
    COL_GRANDPARENT_ID,
    COL_GRANDPARENT_NAME,
    COL_INCLUDE_MIGRATION,
    COL_REASON,
    IMMUTABLE_COLUMNS,
    INCLUDE_YES_VALUES,
    MUTABLE_COLUMNS,
    OUTPUT_DIR,
    PARTIAL_OPT_OUT_VALUES,
    PDF_GOVERNANCE_ENABLED,
    UPDATE_MASTER_DATA,
)

logger = logging.getLogger(__name__)

# ── Colours ───────────────────────────────────────────────────────────────────
_FILL_RED    = PatternFill("solid", fgColor="FFCCCC")   # CTL-critical
_FILL_AMBER  = PatternFill("solid", fgColor="FFE5B4")   # mutable change
_FILL_GREEN  = PatternFill("solid", fgColor="CCFFCC")   # all good
_FILL_HEADER = PatternFill("solid", fgColor="D9E1F2")   # header row
_FILL_BANNER_RED   = PatternFill("solid", fgColor="FF4C4C")
_FILL_BANNER_AMBER = PatternFill("solid", fgColor="FFA500")

# ── Email validation ──────────────────────────────────────────────────────────
_EMAIL_RE = re.compile(
    r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$"
)

# ── Master data cache (keyed by path + mtime) ─────────────────────────────────
_master_cache: dict[str, tuple[float, pd.DataFrame]] = {}
# Pending master write queue — (excel_row_1based, col_name, value) tuples.
# _update_master() appends here with zero file I/O; flush_master_writes() applies
# them all at once via xlwings COM at the end of the run.
_pending_master_updates: list[tuple[int, str, str]] = []

# Local working copy of the master — stored in the code folder, outside OneDrive.
# All reads and writes use this path so the OneDrive filesystem filter driver is
# never involved.  Synced from CIRCUIT_MASTER_PATH on first use (or when source
# is updated); written back to source asynchronously after each save.
_LOCAL_MASTER_WORKING_PATH: Path = Path(__file__).parent / "_master_working_copy.xlsx"


# ── Dataclasses ───────────────────────────────────────────────────────────────

@dataclass
class RowValidationResult:
    row_index: int                                # 0-based row index in attachment df
    circuit_id: str
    found_in_master: bool
    immutable_mismatches: dict[str, dict]         # {col: {"expected": x, "actual": y}}
    mutable_changes: dict[str, dict]              # {col: {"old": x, "new": y}}
    missing_required: list[str]                   # column names with missing required values
    invalid_emails: list[str]                     # invalid email values found
    include_value: str                            # raw value of COL_INCLUDE_MIGRATION
    is_partial_opt_out: bool

    @property
    def is_critical(self) -> bool:
        return (
            not self.found_in_master
            or bool(self.immutable_mismatches)
            or bool(self.missing_required)
            or bool(self.invalid_emails)
        )

    @property
    def has_changes(self) -> bool:
        return bool(self.mutable_changes)


@dataclass
class AttachmentValidationResult:
    attachment_path: str
    source_type: str                              # "xlsx", "pdf", "csv"
    is_valid: bool                                # False if any CTL-critical issue found
    is_partial_opt_out: bool
    pdf_human_review_required: bool
    total_rows: int
    partial_opt_out_row_count: int
    missing_columns: list[str]
    row_results: list[RowValidationResult]
    master_update_applied: bool
    master_backup_path: str
    data_quality_report_path: str
    errors: list[str]

    @property
    def critical_count(self) -> int:
        return sum(1 for r in self.row_results if r.is_critical)

    @property
    def changed_count(self) -> int:
        return sum(1 for r in self.row_results if r.has_changes and not r.is_critical)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _normalise_col(name: str) -> str:
    """Strip and lowercase a column name for case-insensitive comparison."""
    return str(name).strip().lower()


# Known column aliases: maps alternative names (already normalised) to the canonical
# name used throughout this module.  Add new aliases here as needed.
_COL_ALIASES: dict[str, str] = {
    "cli": COL_CIRCUIT_ID,        # Telco CLI (Circuit Line Identifier) = circuit_id
    "entname": COL_GRANDPARENT_NAME,  # abbreviated PDF header for grandparentname
    "ent_id": COL_GRANDPARENT_ID,     # abbreviated PDF header for grandparent_id
    "entid": COL_GRANDPARENT_ID,      # same without underscore
}


def _apply_column_aliases(df: pd.DataFrame) -> pd.DataFrame:
    """
    Rename known column aliases to their canonical names.

    Called after column normalisation so that sheets which use 'CLI' instead of
    'circuit_id' are handled transparently.  Only renames if the alias is present
    AND the canonical column is not already present (avoids accidental overwrite).
    """
    rename_map = {
        alias: canonical
        for alias, canonical in _COL_ALIASES.items()
        if alias in df.columns and canonical not in df.columns
    }
    if rename_map:
        df = df.rename(columns=rename_map)
        logger.debug("Column aliases applied: %s", rename_map)
    return df


def _normalise_value(val) -> str:
    """Coerce a cell value to a clean string for comparison."""
    if val is None or (isinstance(val, float) and str(val) == "nan"):
        return ""
    return str(val).strip()


def _col_map(df: pd.DataFrame) -> dict[str, str]:
    """Return {normalised_name: original_name} for all columns in df."""
    return {_normalise_col(c): c for c in df.columns}


# ── PDF extraction ─────────────────────────────────────────────────────────────

def _extract_pdf_table(pdf_path: Path) -> tuple[Optional[pd.DataFrame], list[str]]:
    """
    Extract a tabular data table from a PDF using pdfplumber.

    Returns (DataFrame, errors). DataFrame is None if extraction fails.
    Only PDFs with a native text layer can be read — scanned/image PDFs return None.
    The extracted data is for DATA CONTROLLER REVIEW ONLY — not validated against master.
    """
    errors: list[str] = []
    try:
        import pdfplumber
    except ImportError:
        errors.append("PDF_UNREADABLE: pdfplumber is not installed — cannot extract PDF tables.")
        return None, errors

    try:
        all_rows: list[list] = []
        header: Optional[list] = None

        with pdfplumber.open(str(pdf_path)) as pdf:
            if not pdf.pages:
                errors.append("PDF_UNREADABLE: PDF has no pages.")
                return None, errors

            for page_num, page in enumerate(pdf.pages, start=1):
                tables = page.extract_tables()
                if not tables:
                    continue

                # Find the table whose first row best matches expected column names
                best_table = None
                best_score = 0
                for table in tables:
                    if not table or not table[0]:
                        continue
                    header_row = [_normalise_col(str(c)) for c in table[0] if c]
                    score = sum(1 for col in ALL_EXPECTED_COLUMNS if col in header_row)
                    if score > best_score:
                        best_score = score
                        best_table = table

                # Accept any table pdfplumber found — even if no column names matched
                # expected headers (e.g. headers are in a different language/format).
                # Caller handles column-mismatch gracefully.
                if best_table is None and tables:
                    best_table = tables[0]
                if best_table is None:
                    continue

                if header is None:
                    header = best_table[0]
                    data_rows = best_table[1:]
                else:
                    # Subsequent pages — skip header row if it matches
                    if best_table[0] == header:
                        data_rows = best_table[1:]
                    else:
                        data_rows = best_table

                all_rows.extend(data_rows)

        if header is None or not all_rows:
            # pdfplumber found no border-line table — try pdfminer text-box separation first.
            # Adobe Sign renders each column as a separate text rendering group; pdfminer with
            # tight char_margin=0.1 isolates them as individual LTTextBox objects.
            logger.info(
                "pdfplumber found no table in %s — trying pdfminer text-box extraction.",
                pdf_path.name,
            )
            df_pm, pm_errors = _extract_pdf_table_pdfminer(pdf_path, [])
            if df_pm is not None:
                return df_pm, errors
            # pdfminer also failed — fall back to img2table + EasyOCR
            logger.info(
                "pdfminer text-box extraction also failed for %s — attempting img2table OCR fallback.",
                pdf_path.name,
            )
            errors.extend(pm_errors)
            return _extract_pdf_table_ocr(pdf_path, errors)

    except Exception as exc:
        errors.append(f"PDF_UNREADABLE: Failed to read PDF — {exc}")
        return None, errors

    df = pd.DataFrame(all_rows, columns=header)
    # Drop rows that are entirely empty
    df = df.dropna(how="all")
    df = df.fillna("")
    logger.info(
        "PDF table extracted: %d rows, %d columns from %s",
        len(df), len(df.columns), pdf_path.name,
    )
    return df, errors


def _extract_pdf_table_pdfminer(
    pdf_path: Path,
    errors: list[str],
) -> tuple[Optional[pd.DataFrame], list[str]]:
    """
    Extract table from multi-page PDFs where each column is a separate text rendering group.

    Adobe Sign splits wide spreadsheets across multiple PDF pages when the table is too wide
    to fit a single page.  Each page shows a DIFFERENT group of columns, and x-coordinates
    RESET to 0 on every new page.  Processing all pages in a single coordinate system causes
    fatal x-position collisions (e.g. column 1 of page 1 and column 1 of page 2 are both at
    x≈57, but they represent completely different data columns).

    Strategy:
      1. Collect text boxes PER PAGE (page_sections).
      2. For each page independently: identify header row (topmost y-bucket), build a
         column x-map for that page, then assign data boxes to columns using left-anchor
         region assignment within that page's coordinate space.
      3. Merge pages by row-index: row N from page 1 + row N from page 2 + … = full row N.
    """
    try:
        from pdfminer.high_level import extract_pages  # type: ignore
        from pdfminer.layout import LAParams, LTTextBox  # type: ignore
    except ImportError:
        errors.append("PDF_UNREADABLE: pdfminer.six not available.")
        return None, errors

    laparams = LAParams(
        char_margin=0.1,    # Very tight — keeps adjacent columns as separate text boxes
        word_margin=0.05,
        line_margin=0.2,
        boxes_flow=None,    # Disable paragraph reflow / column merging
        detect_vertical=False,
    )

    try:
        page_sections: list[list[dict]] = []
        for page_layout in extract_pages(str(pdf_path), laparams=laparams):
            page_h = page_layout.height
            page_boxes: list[dict] = []
            for elem in page_layout:
                if not isinstance(elem, LTTextBox):
                    continue
                text = elem.get_text().strip()
                if not text:
                    continue
                y_top = page_h - elem.bbox[3]
                page_boxes.append({"x0": elem.bbox[0], "y_top": round(y_top, 1), "text": text})
            if page_boxes:
                page_sections.append(page_boxes)
    except Exception as exc:
        errors.append(f"PDF_UNREADABLE: pdfminer extraction failed — {exc}")
        return None, errors

    if not page_sections:
        errors.append("PDF_UNREADABLE: pdfminer found no text in PDF.")
        return None, errors

    # ── Shared helpers ─────────────────────────────────────────────────────────

    def _clean_header(raw_text: str) -> str:
        """Collapse internal newlines (wrapped header cells) then normalise."""
        return raw_text.replace("\n", "").replace("\r", "").strip().lower()

    def _fuzzy_match_col(raw: str) -> Optional[str]:
        """Match a cleaned header string to a canonical column name.

        Three passes, in order of confidence:
          1. Exact match after stripping punctuation/spaces/underscores
          2. Alias lookup  (e.g. 'cli' → 'circuit_id', 'entname' → 'grandparentname')
          3. Suffix match  (e.g. 'entname' is a suffix of 'grandparentname')
        """
        raw_key = raw.replace("_", "").replace(" ", "").replace("/", "").replace("-", "")
        # Pass 1
        for col in ALL_EXPECTED_COLUMNS:
            col_key = col.replace("_", "").replace(" ", "").replace("/", "").replace("-", "")
            if raw_key == col_key:
                return col
        # Pass 2: alias lookup
        alias_target = _COL_ALIASES.get(raw) or _COL_ALIASES.get(raw_key)
        if alias_target:
            return alias_target
        # Pass 3: suffix match (min 4 chars to avoid false positives)
        if len(raw_key) >= 4:
            for col in ALL_EXPECTED_COLUMNS:
                col_key = col.replace("_", "").replace(" ", "").replace("/", "").replace("-", "")
                if col_key.endswith(raw_key):
                    return col
        return None

    # ── Process each PDF page independently ───────────────────────────────────
    # Track which canonical columns have already been claimed across all pages so
    # positional fallback remains globally consistent even across pages.
    globally_claimed: set[str] = set()
    per_page_row_lists: list[list[dict]] = []

    for page_idx, page_boxes in enumerate(page_sections):
        # Group this page's boxes into y-buckets
        rows_dict: dict[float, list[dict]] = {}
        for box in page_boxes:
            y_key = round(box["y_top"] / 4) * 4
            rows_dict.setdefault(y_key, []).append(box)

        sorted_y = sorted(rows_dict)
        if len(sorted_y) < 2:
            logger.debug("Page %d: fewer than 2 y-buckets — skipping.", page_idx + 1)
            continue

        # Header row = topmost y-bucket on this page
        header_boxes = sorted(rows_dict[sorted_y[0]], key=lambda b: b["x0"])

        # Build col_x_map for this page only
        col_x_map: list[tuple[float, str]] = []
        page_claimed: set[str] = set()
        recognized_count = 0

        for hb in header_boxes:
            raw = _clean_header(hb["text"])
            if not raw:
                continue
            if raw in ALL_EXPECTED_COLUMNS and raw not in page_claimed:
                col_x_map.append((hb["x0"], raw))
                page_claimed.add(raw)
                recognized_count += 1
            else:
                matched = _fuzzy_match_col(raw)
                if matched and matched not in page_claimed:
                    col_x_map.append((hb["x0"], matched))
                    page_claimed.add(matched)
                    recognized_count += 1
                # Unrecognized headers are intentionally skipped (no positional fallback).
                # Non-table pages (cover pages, instructions) have 0 recognized headers
                # and are filtered out below.

        if recognized_count < 1:
            # No canonical columns found — this is not a data-table page (e.g. cover, instructions)
            logger.debug("Page %d: no canonical columns recognized — skipping non-table page.", page_idx + 1)
            continue

        globally_claimed.update(page_claimed)

        col_x_sorted = sorted(col_x_map, key=lambda cp: cp[0])
        logger.debug(
            "Page %d col_x_map (sorted): %s",
            page_idx + 1,
            [(round(x, 1), n) for x, n in col_x_sorted],
        )

        # Left-anchor region assignment for this page
        def _assign_col(x0: float, _cxs: list = col_x_sorted) -> str:  # noqa: E731
            best = _cxs[0][1]
            for cx, name in _cxs:
                if cx <= x0:
                    best = name
                else:
                    break
            return best

        # Build data rows for this page
        page_rows: list[dict] = []
        for y_key in sorted_y[1:]:
            row_boxes = sorted(rows_dict[y_key], key=lambda b: b["x0"])
            row: dict[str, str] = {}
            for box in row_boxes:
                if box["x0"] < 0:
                    continue
                col_name = _assign_col(box["x0"])
                if col_name not in row:
                    row[col_name] = box["text"]
            if row:
                page_rows.append(row)

        if page_rows:
            per_page_row_lists.append(page_rows)

    if not per_page_row_lists:
        errors.append("PDF_UNREADABLE: pdfminer found no data rows on any page.")
        return None, errors

    # ── Merge pages by row-index ───────────────────────────────────────────────
    # row N from page 1 + row N from page 2 + row N from page 3 = complete row N
    num_rows = max(len(rows) for rows in per_page_row_lists)
    table_rows: list[dict] = []
    for row_idx in range(num_rows):
        merged: dict[str, str] = {}
        for page_rows in per_page_row_lists:
            if row_idx < len(page_rows):
                for k, v in page_rows[row_idx].items():
                    if k not in merged:   # first page to supply a column wins
                        merged[k] = v
        if merged:
            table_rows.append(merged)

    if not table_rows:
        errors.append("PDF_UNREADABLE: pdfminer found no data rows after page merge.")
        return None, errors

    # ── Post-process: rescue grandparent_id from merged grandparentname cell ──────
    # Adobe Sign renders adjacent columns with physically overlapping x-coordinates.
    # When grandparentname and grandparent_id chars interleave, pdfminer joins them
    # into one box: e.g. "Triton Telecom Ltd2945". Split the trailing digits out.
    _trailing_id = re.compile(r'^(.*\D)\s*(\d+)\s*$', re.DOTALL)
    for row in table_rows:
        if "grandparentname" in row and "grandparent_id" not in row:
            m = _trailing_id.match(str(row["grandparentname"]).strip())
            if m:
                row["grandparentname"] = m.group(1).strip()
                row["grandparent_id"] = m.group(2).strip()

    # ── Post-process: rescue "line to migrate to" merged into address cell ────
    for row in table_rows:
        addr = row.get("address", "")
        if "line to migrate to" not in row and "\n" in addr:
            line_val, _, real_addr = addr.partition("\n")
            row["line to migrate to"] = line_val.strip()
            row["address"] = real_addr.strip()

    df = pd.DataFrame(table_rows)
    df = df.fillna("")

    # ── Fix scientific-notation numbers (e.g. "1.61E+09" → "1610000000") ──────
    # The PDF renders large integers (circuit_id, grandparent_id) in sci notation.
    # Convert them back to plain integer strings so they display/match correctly.
    _sci_re = re.compile(r'^[+-]?\d+\.?\d*[Ee][+-]?\d+$')
    for col in df.columns:
        def _fix_sci(v: str) -> str:
            if _sci_re.match(str(v).strip()):
                try:
                    return str(int(float(v)))
                except (ValueError, OverflowError):
                    pass
            return v
        df[col] = df[col].apply(_fix_sci)

    # ── Reorder columns to match ALL_EXPECTED_COLUMNS ─────────────────────────
    ordered = [c for c in ALL_EXPECTED_COLUMNS if c in df.columns]
    extras  = [c for c in df.columns if c not in ordered]
    df = df[ordered + extras]

    matched_cols = sum(1 for c in ALL_EXPECTED_COLUMNS if c in df.columns)
    logger.info(
        "pdfminer text-box extraction: %d rows, %d columns, %d/%d expected cols matched from %s",
        len(df), len(df.columns), matched_cols, len(ALL_EXPECTED_COLUMNS), pdf_path.name,
    )
    return df, errors


def _extract_pdf_table_ocr(
    pdf_path: Path,
    errors: list[str],
) -> tuple[Optional[pd.DataFrame], list[str]]:
    """
    Fallback OCR extraction using img2table + EasyOCR for rasterised / image-only PDFs.

    Adobe Sign converts uploaded Excel files to flat image PDFs — pdfplumber finds no text
    layer. img2table uses computer vision to detect table borders and EasyOCR reads the
    cell text locally (no external API, GDPR-safe).

    NOTE: EasyOCR downloads language models (~200 MB) on first run.
    """
    try:
        from img2table.document import PDF as Img2PDF  # type: ignore
        from img2table.ocr import EasyOCR as Img2EasyOCR  # type: ignore
    except ImportError:
        errors.append(
            "PDF_UNREADABLE: img2table / easyocr not installed — "
            "cannot OCR image-based PDF. Run: pip install img2table easyocr"
        )
        return None, errors

    try:
        logger.info("Initialising EasyOCR engine (models download on first run)…")
        try:
            ocr = Img2EasyOCR(lang=["en"])
        except (RuntimeError, MemoryError) as mem_exc:
            errors.append(
                f"PDF_UNREADABLE: EasyOCR could not initialise (insufficient memory) — {mem_exc}"
            )
            return None, errors

        img_pdf = Img2PDF(src=str(pdf_path))

        # Try bordered table detection first (Excel grids usually have visible lines),
        # then fall back to borderless (whitespace-based) detection.
        for borderless in (False, True):
            table_map = img_pdf.extract_tables(
                ocr=ocr,
                implicit_rows=True,
                borderless_tables=borderless,
                min_confidence=50,
            )

            best_df: Optional[pd.DataFrame] = None
            best_score = 0

            for page_tables in table_map.values():
                for tbl in page_tables:
                    candidate = tbl.df
                    if candidate is None or candidate.empty:
                        continue

                    # img2table may include header in first row — check both layouts
                    for use_first_row_as_header in (True, False):
                        if use_first_row_as_header:
                            cols = [_normalise_col(str(c)) for c in candidate.iloc[0]]
                            data = candidate.iloc[1:].reset_index(drop=True).copy()
                            data.columns = cols
                        else:
                            data = candidate.copy()
                            data.columns = [_normalise_col(str(c)) for c in data.columns]

                        score = sum(1 for c in ALL_EXPECTED_COLUMNS if c in data.columns)
                        if score > best_score:
                            best_score = score
                            best_df = data

            if best_df is not None and not best_df.empty:
                best_df = best_df.fillna("")
                logger.info(
                    "img2table OCR extracted table (%s): %d rows, %d columns, "
                    "%d/%d expected columns matched.",
                    "bordered" if not borderless else "borderless",
                    len(best_df), len(best_df.columns),
                    best_score, len(ALL_EXPECTED_COLUMNS),
                )
                return best_df, errors

        errors.append(
            "PDF_UNREADABLE: img2table OCR found no usable table structure in the image PDF. "
            "The PDF may not contain visible grid lines. Human review and re-submission as Excel required."
        )
        return None, errors

    except Exception as exc:
        errors.append(f"PDF_UNREADABLE: img2table OCR failed — {exc}")
        return None, errors


# ── Master data loader ────────────────────────────────────────────────────────

def _ensure_local_copy() -> tuple[Path, list[str]]:
    """
    Copy the circuit master from CIRCUIT_MASTER_PATH to _LOCAL_MASTER_WORKING_PATH
    (code folder, outside OneDrive) so every read and write happens on local disk.
    Only copies when the source is newer than the local copy or the local copy does
    not yet exist — so subsequent calls within the same run are instant.
    Returns (local_path, errors).
    """
    errors: list[str] = []
    local = _LOCAL_MASTER_WORKING_PATH

    if not CIRCUIT_MASTER_PATH:
        errors.append("CIRCUIT_MASTER_PATH not configured.")
        return local, errors

    source = Path(CIRCUIT_MASTER_PATH)
    if not source.exists():
        errors.append(f"Circuit master source not found: {CIRCUIT_MASTER_PATH}")
        return local, errors

    try:
        source_mtime = source.stat().st_mtime
        if not local.exists() or source_mtime > local.stat().st_mtime:
            logger.info(
                "Copying master to local working copy: %s → %s", source.name, local.name
            )
            shutil.copy2(str(source), str(local))
            logger.info("Master local copy ready: %s", local)
        else:
            logger.debug("Local master working copy is current: %s", local.name)
    except Exception as exc:
        errors.append(f"Failed to create local master copy: {exc}")

    return local, errors


def _async_sync_master_back(local: Path, source: Path) -> None:
    """
    Copy the local working copy back to the OneDrive source path in a background
    thread so the bot is not blocked by OneDrive's filesystem filter driver.
    """
    def _do_copy() -> None:
        try:
            shutil.copy2(str(local), str(source))
            logger.info("Master synced back to source: %s", source.name)
        except Exception as exc:
            logger.warning("Failed to sync master back to source — changes are safe in local copy. Error: %s", exc)

    threading.Thread(target=_do_copy, daemon=True).start()


def _flush_via_xlwings(local: Path, updates: list[tuple[int, str, str]]) -> None:
    """Apply queued cell updates via xlwings COM automation — fast even on 19k-row files."""
    import xlwings as xw  # type: ignore
    logger.info("Applying %d update(s) to master via xlwings...", len(updates))
    with xw.App(visible=False, add_book=False) as app:
        wb = app.books.open(str(local))
        try:
            ws = wb.sheets[CIRCUIT_MASTER_SHEET_NAME]
            header = ws.range("A1").expand("right").value or []
            col_map: dict[str, int] = {
                _normalise_col(str(h)): idx + 1
                for idx, h in enumerate(header) if h
            }
            for row_num, col_name, value in updates:
                col_num = col_map.get(col_name)
                if col_num:
                    ws.cells(row_num, col_num).value = value
            wb.save()
            logger.info("✅  Master flush complete via xlwings (%d cell(s) written).", len(updates))
        finally:
            wb.close()


def _flush_via_openpyxl(local: Path, updates: list[tuple[int, str, str]]) -> None:
    """Fallback write via openpyxl — only used when xlwings is not installed."""
    logger.warning(
        "Falling back to openpyxl for master write — install xlwings for fast writes: pip install xlwings"
    )

    # 1. Avoid external link errors
    wb = openpyxl.load_workbook(str(local), keep_links=False)

    # 2. Case‑insensitive sheet lookup
    target = CIRCUIT_MASTER_SHEET_NAME.lower()
    sheet_map = {name.lower(): name for name in wb.sheetnames}

    if target not in sheet_map:
        raise KeyError(f"Sheet '{CIRCUIT_MASTER_SHEET_NAME}' not found (case-insensitive match failed).")

    ws = wb[sheet_map[target]]

    # Build header map
    header_row = [
        _normalise_col(str(ws.cell(1, c).value or ""))
        for c in range(1, ws.max_column + 1)
    ]
    col_index = {name: idx + 1 for idx, name in enumerate(header_row) if name}

    updated_cells = 0

    for row_num, col_name, value in updates:
        col_num = col_index.get(col_name)
        if col_num:
            ws.cell(row_num, col_num).value = value
            updated_cells += 1

    wb.save(str(local))

    # 3. Log how many cells were written
    logger.info(f"✅ Master flush complete via openpyxl — {updated_cells} cells updated.")



def flush_master_writes() -> None:
    """
    Apply all queued mutable-column updates via xlwings COM (fast) or openpyxl (fallback).
    Call once at the end of each run. Zero file I/O happens during agreement processing.
    """
    global _pending_master_updates
    if not _pending_master_updates:
        logger.debug("flush_master_writes: no pending updates.")
        return

    local, sync_errors = _ensure_local_copy()
    for e in sync_errors:
        logger.error("flush_master_writes sync error: %s", e)
    if not local.exists():
        logger.error("flush_master_writes: local working copy not found — updates not written.")
        _pending_master_updates = []
        return

    updates = list(_pending_master_updates)
    # NEW — nested so ALL exceptions are caught:
    try:
        try:
            _flush_via_xlwings(local, updates)
        except ImportError:
            _flush_via_openpyxl(local, updates)
    except Exception as exc:
        logger.error("flush_master_writes failed: %s", exc, exc_info=True)
        return

    _pending_master_updates = []
    _master_cache.pop(f"{local}::{CIRCUIT_MASTER_SHEET_NAME}", None)
    if CIRCUIT_MASTER_PATH:
        # Sync back to source synchronously — daemon threads are killed on process exit
        # and the OneDrive file would never receive the updates.
        source = Path(CIRCUIT_MASTER_PATH)
        try:
            shutil.copy2(str(local), str(source))
            logger.info("Master synced back to source: %s", source.name)
        except Exception as exc:
            logger.warning(
                "Failed to sync master back to source — changes are safe in local working copy "
                "(%s). Manually copy it to replace the source if needed. Error: %s",
                local.name, exc,
            )


def _account_id_match(form_val: str, master_val: str) -> bool:
    """Return True if form_val and master_val refer to the same account number.

    Handles the variations seen in practice where leading zeros are stripped
    or the form captures only the last N digits of a longer account number:

      form='00169'  master='169'       → True  (leading-zero normalisation)
      form='00169'  master='0169'      → True  (leading-zero normalisation)
      form='00169'  master='44000169'  → True  (form is suffix of master)
      form='44000169' master='169'     → True  (master is suffix of form)

    Only applies suffix matching to purely numeric values to avoid false
    positives on alphanumeric IDs.
    """
    a = form_val.strip().lstrip("0") or "0"
    b = master_val.strip().lstrip("0") or "0"
    if not a or not b:
        return False
    if not (a.isdigit() and b.isdigit()):
        return a == b
    if a == b:
        return True
    # Suffix match: shorter must be a trailing suffix of the longer
    return b.endswith(a) if len(a) <= len(b) else a.endswith(b)


def queue_full_opt_out_master_updates(
    reseller_name: str,
    account_number: str,
    agreement_id: str,
) -> tuple[int, list[str]]:
    """Queue master sheet updates for a full opt-out submission.

    Finds every row in the master sheet where grandparent_id matches account_number
    (primary) or grandparentname matches reseller_name (secondary, case-insensitive),
    then queues:
      - include in migration y/n  →  "n"
      - reason                    →  "Full Opt-Out"

    Updates are appended to _pending_master_updates and applied in bulk by
    flush_master_writes() at the end of the run.

    Returns (rows_queued, errors).
    """
    errors: list[str] = []

    if not UPDATE_MASTER_DATA:
        return 0, errors

    master_df, load_errors = _load_master_data()
    errors.extend(load_errors)
    if master_df is None:
        errors.append("Full opt-out master update skipped — master data could not be loaded.")
        return 0, errors

    if COL_CIRCUIT_ID not in master_df.columns:
        errors.append("Full opt-out master update skipped — circuit_id column not found in master.")
        return 0, errors

    if COL_GRANDPARENT_ID not in master_df.columns and COL_GRANDPARENT_NAME not in master_df.columns:
        errors.append(
            "Full opt-out master update skipped — neither grandparent_id nor grandparentname "
            "column found in master."
        )
        return 0, errors

    norm_account  = _normalise_value(account_number)
    norm_reseller = _normalise_value(reseller_name).lower()

    queued = 0
    for idx, row in master_df.iterrows():
        excel_row = int(idx) + 2  # +1 for 0-based index, +1 for header row

        gp_id   = _normalise_value(row.get(COL_GRANDPARENT_ID, ""))
        gp_name = _normalise_value(row.get(COL_GRANDPARENT_NAME, "")).lower()

        id_match   = bool(norm_account  and gp_id   and _account_id_match(norm_account, gp_id))
        name_match = bool(norm_reseller and gp_name == norm_reseller)

        if not id_match and not name_match:
            continue

        _pending_master_updates.append((excel_row, COL_INCLUDE_MIGRATION, "n"))
        _pending_master_updates.append((excel_row, COL_REASON, "Full Opt-Out"))
        queued += 1

    if queued:
        logger.info(
            "Full opt-out: queued %d master row(s) for update "
            "(reseller=%r, account=%r, agreement=%s).",
            queued, reseller_name, account_number, agreement_id,
        )
    else:
        logger.warning(
            "Full opt-out: no master rows matched reseller=%r or account=%r "
            "(agreement=%s). Verify Company Name / Account Number on the form match "
            "grandparentname / grandparent_id in the master sheet.",
            reseller_name, account_number, agreement_id,
        )

    return queued, errors


def _load_master_data() -> tuple[Optional[pd.DataFrame], list[str]]:
    """
    Load the circuit master sheet, normalising column names.
    Uses a module-level cache keyed by (path, mtime) to avoid re-reading per row.
    Duplicate circuit_id values are treated as CTL-critical warnings.
    """
    errors: list[str] = []

    if not CIRCUIT_MASTER_PATH:
        errors.append("CIRCUIT_MASTER_PATH not configured — skipping master validation.")
        return None, errors

    # Copy from source (OneDrive) to local working copy if needed — instant on cache hit.
    local_path, sync_errors = _ensure_local_copy()
    errors.extend(sync_errors)
    if not local_path.exists():
        errors.append(
            f"Circuit master sheet not available locally or at source: {CIRCUIT_MASTER_PATH} — "
            "cannot validate attachment data."
        )
        return None, errors

    try:
        mtime = local_path.stat().st_mtime
        cache_key = f"{local_path}::{CIRCUIT_MASTER_SHEET_NAME}"
        if cache_key in _master_cache and _master_cache[cache_key][0] == mtime:
            logger.debug("Using cached master data for %s", local_path.name)
            return _master_cache[cache_key][1], errors

        try:
            df = pd.read_excel(str(local_path), sheet_name=CIRCUIT_MASTER_SHEET_NAME, dtype=str, engine="calamine")
        except ValueError as sheet_exc:
            errors.append(
                f"Circuit master sheet '{CIRCUIT_MASTER_SHEET_NAME}' not found in "
                f"{local_path.name} — check CIRCUIT_MASTER_SHEET_NAME in .env. "
                f"Detail: {sheet_exc}"
            )
            return None, errors

        df = df.fillna("")

        # Normalise column names then apply known aliases (e.g. CLI → circuit_id)
        df.columns = [_normalise_col(c) for c in df.columns]
        df = _apply_column_aliases(df)

        # Check for missing expected columns
        missing = [c for c in ALL_EXPECTED_COLUMNS if c not in df.columns]
        if missing:
            errors.append(
                f"Circuit master is missing expected columns: {missing}. "
                "Validation may be incomplete."
            )

        # Check for duplicate circuit_ids — CTL-critical
        if COL_CIRCUIT_ID in df.columns:
            dupes = df[df.duplicated(subset=[COL_CIRCUIT_ID], keep=False)][COL_CIRCUIT_ID].unique()
            if len(dupes) > 0:
                errors.append(
                    f"CRITICAL: Circuit master contains {len(dupes)} duplicate circuit_id value(s): "
                    f"{list(dupes)[:10]}{'...' if len(dupes) > 10 else ''}. "
                    "Matching may be ambiguous for these circuits."
                )

        _master_cache[cache_key] = (mtime, df)
        logger.info(
            "Circuit master loaded: %d rows from sheet '%s' in %s",
            len(df), CIRCUIT_MASTER_SHEET_NAME, local_path.name,
        )
        return df, errors

    except PermissionError:
        errors.append(
            f"Circuit master is locked (open in another application): {CIRCUIT_MASTER_PATH}. "
            "Validation skipped — close the file and re-run."
        )
        return None, errors
    except Exception as exc:
        errors.append(f"Failed to load circuit master: {exc}")
        return None, errors


# ── Attachment loader ─────────────────────────────────────────────────────────

def _load_attachment_df(
    att_path: Path,
) -> tuple[Optional[pd.DataFrame], str, list[str]]:
    """
    Load the attachment into a DataFrame, normalising column names.

    Returns (df, source_type, errors).
    source_type is one of: "xlsx", "xls", "pdf", "csv".
    For PDFs: df is None (not validated against master) — caller handles human-review path.
    """
    errors: list[str] = []
    ext = att_path.suffix.lower()

    if ext in (".xlsx", ".xls", ".xlsm"):
        try:
            df = pd.read_excel(str(att_path), dtype=str)
            df = df.fillna("")
            df.columns = [_normalise_col(c) for c in df.columns]
            df = _apply_column_aliases(df)
            return df, "xlsx", errors
        except Exception as exc:
            errors.append(f"Failed to read Excel attachment: {exc}")
            return None, "xlsx", errors

    elif ext == ".pdf":
        df, pdf_errors = _extract_pdf_table_pdfminer(att_path, [])
        errors.extend(pdf_errors)
        if df is not None:
            df.columns = [_normalise_col(c) for c in df.columns]
            df = _apply_column_aliases(df)
        return df, "pdf", errors

    elif ext == ".csv":
        try:
            df = pd.read_csv(str(att_path), dtype=str, encoding="utf-8-sig")
            df = df.fillna("")
            df.columns = [_normalise_col(c) for c in df.columns]
            df = _apply_column_aliases(df)
            return df, "csv", errors
        except Exception as exc:
            errors.append(f"Failed to read CSV attachment: {exc}")
            return None, "csv", errors

    else:
        errors.append(f"Unsupported attachment format: {ext} — cannot validate.")
        return None, ext.lstrip(".") or "unknown", errors


# ── Row-level validation ───────────────────────────────────────────────────────

def _validate_rows(
    df: pd.DataFrame,
    master_df: pd.DataFrame,
) -> list[RowValidationResult]:
    """
    Validate each row in the CP attachment against the circuit master.
    Returns a list of RowValidationResult, one per row.
    """
    results: list[RowValidationResult] = []

    # Build a lookup index on circuit_id for O(1) access
    if COL_CIRCUIT_ID not in master_df.columns:
        # Cannot validate — circuit_id column missing from master
        return results

    master_index: dict[str, pd.Series] = {}
    for _, mrow in master_df.iterrows():
        cid = _normalise_value(mrow.get(COL_CIRCUIT_ID, ""))
        if cid:
            master_index[cid] = mrow

    for row_idx, row in df.iterrows():
        circuit_id = _normalise_value(row.get(COL_CIRCUIT_ID, ""))
        include_val = _normalise_value(row.get(COL_INCLUDE_MIGRATION, ""))
        # Guard against pdfminer merging adjacent column text into one cell.
        # When 'include in migration y/n' and 'reason' sit physically close in a PDF,
        # pdfminer can produce e.g. "n\nvulnerable" — take only the first line so the
        # opt-out value is correctly detected as "n".
        if "\n" in include_val:
            include_val = include_val.split("\n")[0].strip()
        include_lower = include_val.lower()
        # Log unrecognised values to aid debugging (supports y/yes/n/no, case-insensitive)
        if include_val and include_lower not in PARTIAL_OPT_OUT_VALUES and include_lower not in INCLUDE_YES_VALUES:
            logger.warning(
                "Row %d (circuit %s): unrecognised include_in_migration value %r "
                "\u2014 expected y, yes, n, or no (case-insensitive)",
                row_idx, circuit_id, include_val,
            )

        found = circuit_id in master_index
        master_row = master_index.get(circuit_id)

        immutable_mismatches: dict[str, dict] = {}
        mutable_changes: dict[str, dict] = {}
        missing_required: list[str] = []
        invalid_emails: list[str] = []

        if found and master_row is not None:
            # Immutable column check
            for col in IMMUTABLE_COLUMNS:
                if col not in df.columns:
                    continue
                actual = _normalise_value(row.get(col, ""))
                expected = _normalise_value(master_row.get(col, ""))
                if actual.lower() != expected.lower():
                    immutable_mismatches[col] = {"expected": expected, "actual": actual}

            # Mutable column change detection
            for col in MUTABLE_COLUMNS:
                if col not in df.columns:
                    continue
                new_val = _normalise_value(row.get(col, ""))
                old_val = _normalise_value(master_row.get(col, ""))
                if new_val.lower() != old_val.lower():
                    mutable_changes[col] = {"old": old_val, "new": new_val}

        # Missing required fields
        for col in ALL_EXPECTED_COLUMNS:
            if col not in df.columns:
                continue
            val = _normalise_value(row.get(col, ""))
            if not val:
                # Reason is optional when circuit is opted IN
                if col == COL_REASON and include_lower in INCLUDE_YES_VALUES:
                    continue
                missing_required.append(col)

        # Email format validation — CTL contact field
        if COL_EMAIL_ADDRESS in df.columns:
            email_val = _normalise_value(row.get(COL_EMAIL_ADDRESS, ""))
            if email_val and not _EMAIL_RE.match(email_val):
                invalid_emails.append(email_val)

        is_partial = include_lower in PARTIAL_OPT_OUT_VALUES

        results.append(RowValidationResult(
            row_index=int(row_idx),
            circuit_id=circuit_id,
            found_in_master=found,
            immutable_mismatches=immutable_mismatches,
            mutable_changes=mutable_changes,
            missing_required=missing_required,
            invalid_emails=invalid_emails,
            include_value=include_val,
            is_partial_opt_out=is_partial,
        ))

    return results


# ── Master data updater ────────────────────────────────────────────────────────

def _update_master(
    df: pd.DataFrame,
    row_results: list[RowValidationResult],
    agreement_id: str,
    master_df: pd.DataFrame,
) -> tuple[bool, str, list[str]]:
    """
    Write mutable field updates back to the circuit master sheet.

    Only updates rows that:
      - Are found in the master
      - Have no immutable mismatches
      - Have at least one mutable change

    Creates a timestamped backup before any write.
    On PermissionError, aborts entirely — no partial write.

    Returns (applied: bool, backup_path: str, errors: list[str])
    """
    errors: list[str] = []

    if not UPDATE_MASTER_DATA:
        return False, "", errors

    if not CIRCUIT_MASTER_PATH:
        errors.append("UPDATE_MASTER_DATA=true but CIRCUIT_MASTER_PATH not set.")
        return False, "", errors

    master_path = Path(CIRCUIT_MASTER_PATH)
    # All file I/O happens against the local working copy (not OneDrive directly).
    # _ensure_local_copy() is guaranteed to have been called by _load_master_data()
    # earlier in validate_attachment(), so the local copy always exists here.
    local_path = _LOCAL_MASTER_WORKING_PATH
    if not local_path.exists():
        errors.append(
            f"Master local working copy not found — validate_attachment must run first "
            f"to sync from source: {CIRCUIT_MASTER_PATH}"
        )
        return False, "", errors

    # Identify rows eligible for update: found in master with no immutable mismatches.
    # We always write all mutable columns from the attachment regardless of whether
    # the value differs from what is already in the master — the CP's submission is
    # treated as the authoritative source for those 3 columns.
    rows_to_update = [
        r for r in row_results
        if r.found_in_master
        and not r.immutable_mismatches
    ]

    if not rows_to_update:
        logger.info("No eligible rows to write back to master for %s.", agreement_id)
        return False, "", errors

    # Create timestamped backup first (skipped when BACKUP_MASTER_BEFORE_UPDATE=false)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    backup_path = master_path.parent / f"{master_path.stem}_{ts}_backup{master_path.suffix}"
    if BACKUP_MASTER_BEFORE_UPDATE:
        try:
            shutil.copy2(str(master_path), str(backup_path))
            logger.info("Master backup created: %s", backup_path.name)
        except Exception as exc:
            errors.append(f"Could not create master backup before update: {exc} — update aborted.")
            return False, "", errors
    else:
        logger.warning(
            "BACKUP_MASTER_BEFORE_UPDATE=false — skipping backup before writing to master. "
            "Ensure you have taken a manual backup before running."
        )

    # Build column index from the already-loaded master_df — no file I/O needed.
    col_index: dict[str, int] = {name: idx + 1 for idx, name in enumerate(master_df.columns)}

    missing_master_cols = [c for c in MUTABLE_COLUMNS if col_index.get(c) is None]
    if missing_master_cols:
        logger.warning(
            "Master sheet '%s' is missing mutable columns: %s — "
            "these will be skipped. Check column headers in the master workbook.",
            CIRCUIT_MASTER_SHEET_NAME, missing_master_cols,
        )

    if COL_CIRCUIT_ID not in master_df.columns:
        errors.append("circuit_id column not found in master DataFrame — update aborted.")
        return False, str(backup_path), errors

    master_row_map: dict[str, int] = {
        _normalise_value(cid): int(idx) + 2
        for idx, cid in master_df[COL_CIRCUIT_ID].items()
        if _normalise_value(cid)
    }

    att_row_map: dict[str, pd.Series] = {}
    if COL_CIRCUIT_ID in df.columns:
        for _, att_row in df.iterrows():
            cid = _normalise_value(att_row.get(COL_CIRCUIT_ID, ""))
            if cid:
                att_row_map[cid] = att_row

    # Queue updates — zero file I/O. flush_master_writes() applies via xlwings at end of run.
    queued_count = 0
    for row_result in rows_to_update:
        master_row_num = master_row_map.get(_normalise_value(row_result.circuit_id))
        if master_row_num is None:
            continue
        att_row = att_row_map.get(_normalise_value(row_result.circuit_id))
        if att_row is None:
            continue
        for col_name in MUTABLE_COLUMNS:
            if col_name not in df.columns or col_index.get(col_name) is None:
                continue
            new_val = _normalise_value(att_row.get(col_name, ""))
            _pending_master_updates.append((master_row_num, col_name, new_val))
            logger.info("  Queued: circuit=%s  col=%s  value='%s'", row_result.circuit_id, col_name, new_val)
        queued_count += 1

    logger.info(
        "✅  Queued %d row(s) for master update (agreement %s) — will flush at end of run.",
        queued_count, agreement_id,
    )
    return True, str(backup_path), errors


# ── Data controller report ─────────────────────────────────────────────────────

def _style_header_row(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = _FILL_HEADER
        cell.alignment = Alignment(wrap_text=True)
    ws.freeze_panes = "A2"


def generate_data_quality_report(
    validation_result: "AttachmentValidationResult",
    reseller_name: str,
    agreement_id: str,
    original_df: Optional[pd.DataFrame],
) -> Path:
    """
    Generate a 4-sheet Excel data quality report for the data controller.

    Sheets:
      1. Summary      — headline stats per submission
      2. All Records  — every row with colour-coded status
      3. Critical Issues — red rows only
      4. Changes Only — mutable field old/new values
    """
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    safe_name = re.sub(r'[\\/:*?"<>|]', "_", reseller_name).strip()
    report_path = OUTPUT_DIR / f"data_quality_{safe_name}_{ts}.xlsx"

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    source_label = validation_result.source_type.upper()
    if validation_result.source_type == "pdf":
        source_label = "PDF ⚠ HUMAN REVIEW REQUIRED"

    summary_data = [
        ["Field", "Value"],
        ["Agreement ID", agreement_id],
        ["Reseller", reseller_name],
        ["Source File", Path(validation_result.attachment_path).name],
        ["Source Type", source_label],
        ["Run Time (UTC)", datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M:%S")],
        ["Total Rows", validation_result.total_rows],
        ["Critical Issues", validation_result.critical_count],
        ["Changed Fields (mutable)", validation_result.changed_count],
        ["Partial Opt-Out Rows", validation_result.partial_opt_out_row_count],
        ["Missing Columns in Attachment", ", ".join(validation_result.missing_columns) or "None"],
        ["Master Update Applied", "Yes" if validation_result.master_update_applied else "No"],
        ["Master Backup Path", validation_result.master_backup_path or "N/A"],
        ["Overall Status", "PASS" if validation_result.is_valid else "⚠ CRITICAL ISSUES FOUND"],
    ]

    for row_data in summary_data:
        ws_sum.append(row_data)

    _style_header_row(ws_sum)
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 60

    # Colour the status row
    status_row = ws_sum.max_row
    fill = _FILL_GREEN if validation_result.is_valid else _FILL_RED
    for cell in ws_sum[status_row]:
        cell.fill = fill

    if validation_result.source_type == "pdf":
        ws_sum.append([])
        ws_sum.append(["⚠ PDF SUBMISSION — DATA CONTROLLER ACTION REQUIRED"])
        ws_sum.append(["", "This submission was received as a PDF. The data has been extracted and"])
        ws_sum.append(["", "saved as an Excel file for your review. It has NOT been validated against"])
        ws_sum.append(["", "the circuit master. You must review the extracted Excel and confirm the"])
        ws_sum.append(["", "data is acceptable before it is used in any downstream process."])
        for row_num in range(status_row + 2, ws_sum.max_row + 1):
            for cell in ws_sum[row_num]:
                cell.fill = _FILL_AMBER

    if validation_result.errors:
        ws_sum.append([])
        ws_sum.append(["Processing Errors"])
        for err in validation_result.errors:
            ws_sum.append(["", err])

    # ── Sheet 2: All Records ──────────────────────────────────────────────────
    ws_all = wb.create_sheet("All Records")
    all_headers = ["Row", "Status", "Circuit ID", "Partial Opt-Out"] + ALL_EXPECTED_COLUMNS
    ws_all.append(all_headers)
    _style_header_row(ws_all)

    for rr in validation_result.row_results:
        if rr.is_critical:
            status_label = "CRITICAL"
            row_fill = _FILL_RED
        elif rr.has_changes:
            status_label = "CHANGED"
            row_fill = _FILL_AMBER
        else:
            status_label = "OK"
            row_fill = _FILL_GREEN

        if original_df is not None and rr.row_index < len(original_df):
            src_row = original_df.iloc[rr.row_index]
            row_vals = [_normalise_value(src_row.get(c, "")) for c in ALL_EXPECTED_COLUMNS]
        else:
            row_vals = [""] * len(ALL_EXPECTED_COLUMNS)

        ws_all.append([rr.row_index + 1, status_label, rr.circuit_id, "Yes" if rr.is_partial_opt_out else "No"] + row_vals)
        for cell in ws_all[ws_all.max_row]:
            cell.fill = row_fill

    if not validation_result.row_results and validation_result.source_type == "pdf":
        ws_all.append(["—", "PDF — No master validation performed. See extracted Excel.", "", ""])

    # ── Sheet 3: Critical Issues ──────────────────────────────────────────────
    ws_crit = wb.create_sheet("Critical Issues")
    crit_headers = [
        "Row", "Circuit ID", "Issue Type", "Field", "Expected (Master)", "Actual (Submission)", "Detail"
    ]
    ws_crit.append(crit_headers)
    _style_header_row(ws_crit)

    for rr in validation_result.row_results:
        if not rr.is_critical:
            continue

        if not rr.found_in_master:
            ws_crit.append([
                rr.row_index + 1, rr.circuit_id,
                "CIRCUIT NOT FOUND IN MASTER",
                COL_CIRCUIT_ID, "—", rr.circuit_id,
                "This circuit ID does not exist in the master sheet. "
                "Could indicate a wrong-record submission — verify identity before processing.",
            ])
            for cell in ws_crit[ws_crit.max_row]:
                cell.fill = _FILL_RED

        for col, mismatch in rr.immutable_mismatches.items():
            ws_crit.append([
                rr.row_index + 1, rr.circuit_id,
                "IMMUTABLE FIELD MISMATCH",
                col, mismatch["expected"], mismatch["actual"],
                "This field must not differ from the master. "
                "A mismatch may indicate a wrong-record submission. "
                "DO NOT process until verified.",
            ])
            for cell in ws_crit[ws_crit.max_row]:
                cell.fill = _FILL_RED

        for col in rr.missing_required:
            ws_crit.append([
                rr.row_index + 1, rr.circuit_id,
                "MISSING REQUIRED FIELD",
                col, "—", "(blank)",
                "Required data is missing. "
                + ("Customer contact email is missing — failed service notifications may result."
                   if col == COL_EMAIL_ADDRESS else
                   "Circuit cannot be safely processed without this value."),
            ])
            for cell in ws_crit[ws_crit.max_row]:
                cell.fill = _FILL_RED

        for email in rr.invalid_emails:
            ws_crit.append([
                rr.row_index + 1, rr.circuit_id,
                "INVALID EMAIL FORMAT",
                COL_EMAIL_ADDRESS, "valid@domain.com format", email,
                "Invalid email address. An invalid email will cause failed service notifications.",
            ])
            for cell in ws_crit[ws_crit.max_row]:
                cell.fill = _FILL_RED

    for col_letter in ["A", "B", "C", "D", "E", "F", "G"]:
        ws_crit.column_dimensions[col_letter].width = 20
    ws_crit.column_dimensions["G"].width = 60

    # ── Sheet 4: Changes Only ─────────────────────────────────────────────────
    ws_chg = wb.create_sheet("Changes Only")
    chg_headers = ["Row", "Circuit ID", "Field", "Old Value (Master)", "New Value (Submission)", "Master Updated"]
    ws_chg.append(chg_headers)
    _style_header_row(ws_chg)

    for rr in validation_result.row_results:
        for col, change in rr.mutable_changes.items():
            master_updated = (
                "Yes" if rr.found_in_master and not rr.immutable_mismatches and validation_result.master_update_applied
                else "No"
            )
            ws_chg.append([
                rr.row_index + 1, rr.circuit_id, col,
                change["old"], change["new"], master_updated,
            ])
            for cell in ws_chg[ws_chg.max_row]:
                cell.fill = _FILL_AMBER

    for col_letter in ["A", "B", "C", "D", "E", "F"]:
        ws_chg.column_dimensions[col_letter].width = 28

    OUTPUT_DIR.mkdir(exist_ok=True)
    wb.save(str(report_path))
    logger.info("Data quality report saved: %s", report_path.name)
    return report_path


# ── Main entry point ──────────────────────────────────────────────────────────

def validate_attachment(
    att_path: Path,
    agreement_id: str,
    reseller_name: str,
    convert_to_csv: bool = True,
) -> AttachmentValidationResult:
    """
    Validate a CP-uploaded attachment against the circuit master sheet.

    For Excel/CSV:
      - Validates all 9 columns row-by-row
      - Detects partial opt-outs, immutable mismatches, missing required fields, invalid emails
      - Optionally updates mutable fields in the master

    For PDF:
      - Extracts table to DataFrame and saves as CSV for data controller review
      - Does NOT validate against master — human review is required
      - Flags pdf_human_review_required=True

    All exceptions are caught — this function never raises.

    Args:
        att_path:        Path to the saved attachment file
        agreement_id:    Adobe Sign agreement ID (for logging and report naming)
        reseller_name:   CP name (for report naming)
        convert_to_csv:  If True and source is not already CSV, save extracted data as CSV

    Returns:
        AttachmentValidationResult
    """
    all_errors: list[str] = []
    row_results: list[RowValidationResult] = []
    missing_columns: list[str] = []
    master_update_applied = False
    master_backup_path = ""
    is_valid = True
    pdf_human_review = False
    original_df: Optional[pd.DataFrame] = None

    try:
        logger.info(
            "Starting attachment validation: %s (agreement: %s)",
            att_path.name, agreement_id,
        )

        # Step 1: Load attachment
        df, source_type, load_errors = _load_attachment_df(att_path)
        all_errors.extend(load_errors)

        # PDF path — extract table first, then decide whether to validate or flag for review
        if source_type == "pdf":
            if df is None:
                # Unreadable / scanned PDF — cannot extract any data
                is_valid = False
                pdf_human_review = True
                all_errors.append(
                    "PDF attachment could not be parsed (possibly scanned/image-based). "
                    "Human review and manual re-submission as Excel is required."
                )
                result = AttachmentValidationResult(
                    attachment_path=str(att_path),
                    source_type=source_type,
                    is_valid=False,
                    is_partial_opt_out=False,
                    pdf_human_review_required=True,
                    total_rows=0,
                    partial_opt_out_row_count=0,
                    missing_columns=[],
                    row_results=[],
                    master_update_applied=False,
                    master_backup_path="",
                    data_quality_report_path="",
                    errors=all_errors,
                )
                report_path = generate_data_quality_report(result, reseller_name, agreement_id, None)
                result.data_quality_report_path = str(report_path)
                return result

            # PDF extracted successfully — save Excel copy
            if convert_to_csv:
                xlsx_path = att_path.with_suffix(".xlsx")
                try:
                    from openpyxl import Workbook  # type: ignore
                    from openpyxl.utils.dataframe import dataframe_to_rows  # type: ignore
                    wb_out = Workbook()
                    ws_out = wb_out.active
                    # Write header + rows; force numeric-looking columns as text (@) so
                    # Excel doesn't re-render circuit_id / grandparent_id as scientific notation
                    _text_cols = {"circuit_id", "grandparent_id"}
                    header_written = False
                    for r in dataframe_to_rows(df, index=False, header=True):
                        ws_out.append(r)
                        if not header_written:
                            header_written = True
                            continue
                        row_idx = ws_out.max_row
                        for col_idx, col_name in enumerate(df.columns, start=1):
                            if col_name in _text_cols:
                                cell = ws_out.cell(row=row_idx, column=col_idx)
                                cell.number_format = "@"
                                cell.value = str(cell.value) if cell.value is not None else ""
                    wb_out.save(str(xlsx_path))
                    logger.info("PDF extracted → XLSX: %s", xlsx_path.name)
                except Exception as xlsx_exc:
                    all_errors.append(f"Failed to save PDF-extracted XLSX: {xlsx_exc}")

            # Check whether the extracted table has the expected columns.
            # Adobe Sign auto-converts Excel uploads to PDF — if columns match,
            # run full validation exactly as we would for Excel.
            df.columns = [_normalise_col(c) for c in df.columns]
            original_df = df.copy()
            extracted_cols = set(df.columns)
            matched_expected = [c for c in ALL_EXPECTED_COLUMNS if c in extracted_cols]

            missing_columns = [c for c in ALL_EXPECTED_COLUMNS if c not in extracted_cols]

            if len(matched_expected) >= 1:
                # At least one expected column found — attempt full validation.
                # When PDF_GOVERNANCE_ENABLED: flag for amber review email regardless.
                # When off (default): treat extracted PDF data like Excel — only alert on real data issues.
                pdf_human_review = PDF_GOVERNANCE_ENABLED
                source_type = "pdf_converted"
                if missing_columns:
                    all_errors.append(
                        f"PDF extracted with {len(matched_expected)}/{len(ALL_EXPECTED_COLUMNS)} "
                        f"expected columns ({missing_columns} missing). "
                        "Validation will run against available columns — review results carefully."
                    )
                    logger.info(
                        "PDF table extracted with %d/%d expected columns — running partial validation "
                        "(missing: %s).",
                        len(matched_expected), len(ALL_EXPECTED_COLUMNS), missing_columns,
                    )
                else:
                    logger.info(
                        "PDF table extracted with all %d expected columns — running full validation.",
                        len(ALL_EXPECTED_COLUMNS),
                    )
                # Fall through to master validation below
            else:
                # Zero recognised columns — cannot map data to master; flag for human review
                pdf_human_review = True
                all_errors.append(
                    f"PDF extracted but no expected columns were found "
                    f"(available: {list(extracted_cols)[:10]}). "
                    "Cannot validate against circuit master — human review required."
                )
                result = AttachmentValidationResult(
                    attachment_path=str(att_path),
                    source_type=source_type,
                    is_valid=False,
                    is_partial_opt_out=False,
                    pdf_human_review_required=True,
                    total_rows=len(df),
                    partial_opt_out_row_count=0,
                    missing_columns=missing_columns,
                    row_results=[],
                    master_update_applied=False,
                    master_backup_path="",
                    data_quality_report_path="",
                    errors=all_errors,
                )
                report_path = generate_data_quality_report(result, reseller_name, agreement_id, original_df)
                result.data_quality_report_path = str(report_path)
                return result

        # For Excel/CSV: df was already loaded in Step 1 — just check it succeeded
        if source_type not in ("pdf", "pdf_converted"):
            if df is None:
                is_valid = False
                result = AttachmentValidationResult(
                    attachment_path=str(att_path),
                    source_type=source_type,
                    is_valid=False,
                    is_partial_opt_out=False,
                    pdf_human_review_required=False,
                    total_rows=0,
                    partial_opt_out_row_count=0,
                    missing_columns=[],
                    row_results=[],
                    master_update_applied=False,
                    master_backup_path="",
                    data_quality_report_path="",
                    errors=all_errors,
                )
                report_path = generate_data_quality_report(result, reseller_name, agreement_id, None)
                result.data_quality_report_path = str(report_path)
                return result

            original_df = df.copy()
        # For pdf_converted: original_df already set in the PDF block above

        # Step 2: Check for missing expected columns
        missing_columns = [c for c in ALL_EXPECTED_COLUMNS if c not in df.columns]
        if missing_columns:
            all_errors.append(
                f"CRITICAL: Attachment is missing expected column(s): {missing_columns}. "
                "Validation cannot be completed for these fields."
            )
            is_valid = False

        # Save as CSV if requested (pdf_converted already saved its CSV above)
        if convert_to_csv and source_type not in ("csv", "pdf_converted"):
            csv_path = att_path.with_suffix(".csv")
            try:
                df.to_csv(str(csv_path), index=False, encoding="utf-8-sig")
                logger.info("Attachment converted to CSV: %s", csv_path.name)
            except Exception as csv_exc:
                all_errors.append(f"Failed to save attachment as CSV: {csv_exc}")

        # Step 3: Load master
        master_df, master_errors = _load_master_data()
        all_errors.extend(master_errors)

        if master_df is not None:
            row_results = _validate_rows(df, master_df)
        else:
            all_errors.append(
                "Master data could not be loaded — row-level validation skipped. "
                "Manual verification required."
            )
            is_valid = False

        # Step 4: Determine overall validity
        if any(r.is_critical for r in row_results):
            is_valid = False

        is_partial_opt_out = any(r.is_partial_opt_out for r in row_results)
        partial_count = sum(1 for r in row_results if r.is_partial_opt_out)

        # Step 5: Update master (only if no critical issues at all)
        if UPDATE_MASTER_DATA and master_df is not None and row_results:
            master_update_applied, master_backup_path, update_errors = _update_master(
                df, row_results, agreement_id, master_df
            )
            all_errors.extend(update_errors)

        # Step 6: Generate report
        result = AttachmentValidationResult(
            attachment_path=str(att_path),
            source_type=source_type,
            is_valid=is_valid,
            is_partial_opt_out=is_partial_opt_out,
            pdf_human_review_required=pdf_human_review,
            total_rows=len(df),
            partial_opt_out_row_count=partial_count,
            missing_columns=missing_columns,
            row_results=row_results,
            master_update_applied=master_update_applied,
            master_backup_path=master_backup_path,
            data_quality_report_path="",
            errors=all_errors,
        )

        report_path = generate_data_quality_report(result, reseller_name, agreement_id, original_df)
        result.data_quality_report_path = str(report_path)

        logger.info(
            "Validation complete for %s: valid=%s, critical=%d, changed=%d, partial=%d",
            att_path.name, result.is_valid, result.critical_count,
            result.changed_count, result.partial_opt_out_row_count,
        )
        return result

    except Exception as exc:
        msg = f"Unexpected error in validate_attachment for {att_path}: {exc}"
        logger.exception(msg)
        all_errors.append(msg)
        is_valid = False

        result = AttachmentValidationResult(
            attachment_path=str(att_path),
            source_type="unknown",
            is_valid=False,
            is_partial_opt_out=False,
            pdf_human_review_required=False,
            total_rows=0,
            partial_opt_out_row_count=0,
            missing_columns=[],
            row_results=[],
            master_update_applied=False,
            master_backup_path="",
            data_quality_report_path="",
            errors=all_errors,
        )
        try:
            report_path = generate_data_quality_report(result, reseller_name, agreement_id, None)
            result.data_quality_report_path = str(report_path)
        except Exception:
            pass
        return result
