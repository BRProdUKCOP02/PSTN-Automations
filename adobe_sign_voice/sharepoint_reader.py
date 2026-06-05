"""
sharepoint_reader.py — Reads the master data Excel file from a UNC/mapped path.

Expected columns (case-insensitive, extra columns ignored):
    partner_name        — Display name of the Channel Partner
    partner_email       — Email address to send the Adobe Sign agreement to
    account_ref         — Gamma account reference number
    circuit_lines       — (optional) comma-separated list of circuit/line IDs
    any additional columns are passed through as extra merge fields

The reader returns a list of PartnerRecord dicts ready for bulk_sender.py.
"""
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from config import MASTER_DATA_PATH

logger = logging.getLogger(__name__)

# Required columns — these MUST exist (after lower-casing)
REQUIRED_COLUMNS = {"partner_name", "partner_email"}


def _normalise_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Lower-case and strip whitespace from all column names."""
    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]
    return df


def _validate_row(row: dict, row_index: int) -> list[str]:
    """Return a list of validation error messages for a single row."""
    errors = []
    if not str(row.get("partner_name", "")).strip():
        errors.append(f"Row {row_index}: partner_name is blank")
    email = str(row.get("partner_email", "")).strip()
    if not email:
        errors.append(f"Row {row_index}: partner_email is blank")
    elif "@" not in email or "." not in email.split("@")[-1]:
        errors.append(f"Row {row_index}: partner_email '{email}' looks invalid")
    return errors


def load_master_data(path: str = MASTER_DATA_PATH) -> list[dict[str, Any]]:
    """
    Load and validate the master data file.

    Returns a list of dicts, one per partner.
    Raises ValueError if required columns are missing or any rows fail validation.
    Skips fully empty rows silently.
    """
    file_path = Path(path)
    if not file_path.exists():
        raise FileNotFoundError(
            f"Master data file not found: {file_path}\n"
            f"Check MASTER_DATA_PATH in .env"
        )

    logger.info("Loading master data from: %s", file_path)
    df = pd.read_excel(file_path, dtype=str)
    df = _normalise_columns(df)

    # Check required columns exist
    missing = REQUIRED_COLUMNS - set(df.columns)
    if missing:
        raise ValueError(
            f"Master data is missing required columns: {missing}\n"
            f"Found columns: {list(df.columns)}"
        )

    # Drop fully empty rows
    df.dropna(how="all", inplace=True)
    df.fillna("", inplace=True)

    # Validate and collect records
    records: list[dict[str, Any]] = []
    all_errors: list[str] = []

    for idx, row in df.iterrows():
        row_dict = {k: str(v).strip() for k, v in row.items()}
        # Skip rows where all required fields are blank (e.g. trailing whitespace rows)
        if not any(row_dict.get(c) for c in REQUIRED_COLUMNS):
            continue
        errors = _validate_row(row_dict, idx + 2)  # +2 = 1-based + header row
        if errors:
            all_errors.extend(errors)
        else:
            row_dict["_excel_row"] = idx + 2  # 1-based row number in Excel (header = row 1)
            records.append(row_dict)

    if all_errors:
        raise ValueError(
            f"Master data validation failed with {len(all_errors)} error(s):\n"
            + "\n".join(f"  • {e}" for e in all_errors)
        )

    logger.info("Loaded %d partner records from master data.", len(records))
    return records


def update_excel_after_send(
    excel_row: int,
    agreement_id: str,
    path: str = MASTER_DATA_PATH,
) -> None:
    """
    Write agreement_id and date_sent back to the Excel file for a specific row.
    Uses openpyxl to preserve existing formatting and other cell values.
    """
    file_path = Path(path)
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Build column index from header row
    headers = {str(cell.value).strip().lower().replace(" ", "_"): cell.column
               for cell in ws[1] if cell.value}

    date_sent_col = headers.get("date_sent")
    agreement_id_col = headers.get("agreement_id")

    if date_sent_col:
        ws.cell(row=excel_row, column=date_sent_col).value = (
            datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
        )
    if agreement_id_col:
        ws.cell(row=excel_row, column=agreement_id_col).value = agreement_id

    wb.save(file_path)
    logger.info("Excel updated: row %d — agreement_id and date_sent written.", excel_row)


def try_update_excel_after_send(excel_row: int, agreement_id: str, path: str = MASTER_DATA_PATH) -> None:
    """Calls update_excel_after_send but logs a warning rather than raising on failure."""
    try:
        update_excel_after_send(excel_row, agreement_id, path)
    except Exception as exc:
        logger.warning(
            "Could not write back to Excel (row %d): %s — "
            "manually update agreement_id and date_sent.",
            excel_row, exc
        )


def update_excel_after_completion(agreement_id: str, path: str = MASTER_DATA_PATH) -> None:
    """
    Write date_received back to the Excel row that matches the given agreement_id.
    Searches the agreement_id column to find the correct row.
    """
    file_path = Path(path)
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    headers = {str(cell.value).strip().lower().replace(" ", "_"): cell.column
               for cell in ws[1] if cell.value}

    agreement_id_col = headers.get("agreement_id")
    date_received_col = headers.get("date_received")

    if not agreement_id_col:
        logger.warning("No 'Agreement ID' column found in Excel — skipping date_received write-back.")
        return

    # Find the row with matching agreement_id
    target_row = None
    for row in ws.iter_rows(min_row=2, values_only=False):
        cell = row[agreement_id_col - 1]
        if str(cell.value).strip() == agreement_id:
            target_row = cell.row
            break

    if target_row is None:
        logger.warning("Agreement ID %s not found in Excel — skipping date_received write-back.", agreement_id)
        return

    if date_received_col:
        ws.cell(row=target_row, column=date_received_col).value = (
            datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
        )

    wb.save(file_path)
    logger.info("Excel updated: row %d — date_received written for agreement %s.", target_row, agreement_id)


def try_update_excel_after_completion(agreement_id: str, path: str = MASTER_DATA_PATH) -> None:
    """Calls update_excel_after_completion but logs a warning rather than raising on failure."""
    try:
        update_excel_after_completion(agreement_id, path)
    except Exception as exc:
        logger.warning(
            "Could not write date_received to Excel for %s: %s — update manually.",
            agreement_id, exc
        )


def build_merge_fields(partner: dict[str, Any]) -> list[dict]:
    """
    Convert a partner record into the Adobe Sign mergeFieldInfo list format.

    Pre-fills all fields from master data as read-only locked values.
    The partner will complete non-pre-filled fields interactively in the form.

    Fields mapped:
        partner_name  → merge field name: partner_name
        account_ref   → merge field name: account_ref
        circuit_lines → merge field name: circuit_lines
        + any other columns present
    """
    # Fields that map directly to Adobe Sign merge field names in the template
    FIELD_MAPPING = {
        "partner_name": "partner_name",
        "account_ref": "account_ref",
        "circuit_lines": "circuit_lines",
    }
    merge_fields = []
    for col, field_name in FIELD_MAPPING.items():
        value = partner.get(col, "")
        if value:
            merge_fields.append({"fieldName": field_name, "defaultValue": value})

    # Include any extra columns not in FIELD_MAPPING (if template supports them)
    reserved = {"partner_email"} | set(FIELD_MAPPING.keys())
    for col, value in partner.items():
        if col not in reserved and value:
            merge_fields.append({"fieldName": col, "defaultValue": value})

    return merge_fields


def write_to_audit_tab(form_data: dict, agreement_id: str, path: str = MASTER_DATA_PATH) -> None:
    """
    Append a processed widget submission to the 'Audit' tab in the master Excel file.
    Creates the Audit sheet if it does not already exist.

    Columns written:
        agreement_id, signer_name, signer_email, account_ref, migration_status,
        opt_out_full, rwp_remain, confirm_amendments, processed_at
    """
    file_path = Path(path)
    wb = openpyxl.load_workbook(file_path)

    if "Audit" not in wb.sheetnames:
        ws_audit = wb.create_sheet("Audit")
        ws_audit.append([
            "Agreement ID", "Signer Name", "Signer Email", "Account Ref",
            "Migration Status", "Full Opt-Out", "RWP Remain",
            "Confirm Amendments", "Processed At",
        ])
    else:
        ws_audit = wb["Audit"]

    ws_audit.append([
        agreement_id,
        form_data.get("signer_name", ""),
        form_data.get("signer_email", ""),
        form_data.get("account_ref", ""),
        form_data.get("migration_status", ""),
        "Yes" if form_data.get("opt_out_full") else "No",
        "Yes" if form_data.get("rwp_remain") else "No",
        "Yes" if form_data.get("confirm_amendments") else "No",
        form_data.get("processed_at", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")),
    ])

    wb.save(file_path)
    logger.info("Audit tab updated for agreement %s.", agreement_id)


def try_write_to_audit_tab(form_data: dict, agreement_id: str, path: str = MASTER_DATA_PATH) -> None:
    """Calls write_to_audit_tab but logs a warning rather than raising on failure."""
    try:
        write_to_audit_tab(form_data, agreement_id, path)
    except Exception as exc:
        logger.warning(
            "Could not write to Audit tab for %s: %s — update manually.",
            agreement_id, exc
        )
