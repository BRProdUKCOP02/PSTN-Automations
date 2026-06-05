"""
widget_monitor.py — Monitors Adobe Sign widget submissions and processes new agreements.

CPs access the same Adobe Sign webform via a single link. This script polls the
widget for completed submissions, processes each new one using the existing
response_processor logic, writes results to the Audit tab in the master Excel,
and sends a summary report to the Professional Services mailbox after each run.

State is tracked in input/widget_state.json (dict keyed by agreement_id) to
ensure each submission is processed exactly once.

Usage:
    python widget_monitor.py

Scheduled task:
    python orchestrator.py --widget-monitor
"""
import base64
import json
import logging
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import msal
import openpyxl
from openpyxl.styles import Font
import requests

from adobe_sign_client import AdobeSignClient, AdobeSignError
from config import (
    ADOBE_SIGN_WIDGET_ID,
    CHASER_EMAIL,
    CIRCUIT_MASTER_PATH,
    FIELD_CONFIRM_AMENDMENTS,
    FIELD_CONFIRM_RWP_REMAIN,
    FIELD_OPT_OUT_FULL,
    GITHUB_REPO,
    GITHUB_TOKEN,
    GRAPH_CLIENT_ID,
    GRAPH_CLIENT_SECRET,
    GRAPH_SENDER_MAILBOX,
    GRAPH_TENANT_ID,
    LOGS_DIR,
    PS_SUMMARY_EMAIL,
    UPDATE_MASTER_DATA,
)
from attachment_validator import flush_master_writes
from git_backup import backup_master_to_github
from response_processor import process_signed_agreement
from sharepoint_reader import try_write_to_audit_tab

# ── Logging ───────────────────────────────────────────────────────────────────
LOGS_DIR.mkdir(exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(name)-25s  %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(
            LOGS_DIR / f"widget_monitor_{datetime.now(timezone.utc).strftime('%Y%m%d')}.log",
            encoding="utf-8",
        ),
    ],
)
logger = logging.getLogger("widget_monitor")


def _parse_email_list(email_config: str) -> list[dict]:
    """Convert comma-separated emails to Graph API recipient format.
    
    Args:
        email_config: Single email or comma-separated list of emails
        
    Returns:
        List of recipient dicts in Graph API format: [{"emailAddress": {"address": "..."}}, ...]
    """
    if not email_config:
        return []
    return [
        {"emailAddress": {"address": addr.strip()}}
        for addr in email_config.split(",")
        if addr.strip()
    ]


# ── State ─────────────────────────────────────────────────────────────────────
_WIDGET_STATE_FILE = Path(__file__).parent / "input" / "widget_state.json"

# Statuses that mean the agreement is fully completed
_COMPLETED_STATUSES = {"SIGNED", "APPROVED", "ACCEPTED", "FORM_FILLED", "DELIVERED"}


def _load_widget_state() -> dict:
    """Load widget state dict: {agreement_id: {signer_name, processed_at, ...}}"""
    if _WIDGET_STATE_FILE.exists():
        with open(_WIDGET_STATE_FILE, encoding="utf-8-sig") as fh:
            return json.load(fh)
    return {}


def _save_widget_state(state: dict) -> None:
    _WIDGET_STATE_FILE.parent.mkdir(exist_ok=True)
    with open(_WIDGET_STATE_FILE, "w", encoding="utf-8") as fh:
        json.dump(state, fh, indent=2, ensure_ascii=False)


def _get_signer_name(agreement: dict) -> str:
    """Extract the signer's display name from the agreement participantSetsInfo."""
    try:
        sets = agreement.get("participantSetsInfo", [])
        if sets:
            members = sets[0].get("memberInfos", [])
            if members:
                return members[0].get("name", "") or members[0].get("email", "unknown")
    except Exception:
        pass
    return agreement.get("name", "unknown")


def _get_signer_email(agreement: dict) -> str:
    """Extract the signer's email from the agreement participantSetsInfo."""
    try:
        sets = agreement.get("participantSetsInfo", [])
        if sets:
            members = sets[0].get("memberInfos", [])
            if members:
                return members[0].get("email", "")
    except Exception:
        pass
    return ""


# ── PS Summary Email ──────────────────────────────────────────────────────────

def _get_graph_token() -> str:
    app = msal.ConfidentialClientApplication(
        GRAPH_CLIENT_ID,
        authority=f"https://login.microsoftonline.com/{GRAPH_TENANT_ID}",
        client_credential=GRAPH_CLIENT_SECRET,
    )
    result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    if "access_token" not in result:
        raise RuntimeError(
            f"Graph API authentication failed: {result.get('error_description')}"
        )
    return result["access_token"]


def _send_ps_summary(
    processed: list[dict],
    errors: list[dict],
    skipped: int,
) -> None:
    """Send a summary email to the PS mailbox after each widget monitor run."""
    if not PS_SUMMARY_EMAIL:
        logger.warning("PS_SUMMARY_EMAIL not set — skipping summary email.")
        return

    run_time = datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC")
    total = len(processed) + len(errors)

    rows_processed = "".join(
        f"<tr><td style='padding:4px 12px 4px 0'>{r['signer_name']}</td>"
        f"<td style='padding:4px 12px 4px 0'>{r['signer_email']}</td>"
        f"<td style='padding:4px 12px 4px 0'>{r['agreement_id']}</td>"
        f"<td style='padding:4px 0'>{'Full Opt-Out' if r.get('opt_out_full') else 'Acknowledged'}</td></tr>"
        for r in processed
    ) or "<tr><td colspan='4'>None</td></tr>"

    rows_errors = "".join(
        f"<tr><td style='padding:4px 12px 4px 0'>{e['agreement_id']}</td>"
        f"<td style='padding:4px 0'>{e['error']}</td></tr>"
        for e in errors
    ) or "<tr><td colspan='2'>None</td></tr>"

    # Data quality summary rows
    dq_items = [
        r for r in processed
        if r.get("data_quality_issues") or r.get("partial_opt_out") or r.get("pdf_human_review")
    ]
    rows_dq = "".join(
        f"<tr style='background:{'#FFCCCC' if r.get('data_quality_issues') else '#FFE5B4'};'>"
        f"<td style='padding:4px 8px'>{r['agreement_id']}</td>"
        f"<td style='padding:4px 8px'>{r.get('reseller_name', r['signer_name'])}</td>"
        f"<td style='padding:4px 8px'>{'⚠ Yes' if r.get('data_quality_issues') else 'No'}</td>"
        f"<td style='padding:4px 8px'>{r.get('changed_fields', '')}</td>"
        f"<td style='padding:4px 8px'>{'Yes' if r.get('partial_opt_out') else 'No'}</td>"
        f"<td style='padding:4px 8px'>{'⚠ Yes' if r.get('pdf_human_review') else 'No'}</td>"
        "</tr>"
        for r in dq_items
    ) or "<tr><td colspan='6' style='padding:4px'>None</td></tr>"

    html_body = (
        '<html><body style="font-family:Arial,sans-serif;font-size:14px;color:#333;">'
        f"<p><strong>PSTN Voice Migration — Widget Monitor Run Summary</strong><br>"
        f"Run time: {run_time}</p>"
        f"<p>Total submissions seen: <strong>{total + skipped}</strong> &nbsp;|&nbsp; "
        f"Newly processed: <strong>{len(processed)}</strong> &nbsp;|&nbsp; "
        f"Skipped (already done): <strong>{skipped}</strong> &nbsp;|&nbsp; "
        f"Errors: <strong>{len(errors)}</strong></p>"
        "<h3 style='margin-bottom:4px;'>Processed Submissions</h3>"
        '<table style="border-collapse:collapse;margin-bottom:16px;">'
        "<tr style='background:#f0f0f0'>"
        "<th style='padding:4px 12px 4px 0;text-align:left'>Signer Name</th>"
        "<th style='padding:4px 12px 4px 0;text-align:left'>Email</th>"
        "<th style='padding:4px 12px 4px 0;text-align:left'>Agreement ID</th>"
        "<th style='padding:4px 0;text-align:left'>Selection</th></tr>"
        f"{rows_processed}</table>"
        "<h3 style='margin-bottom:4px;'>Errors</h3>"
        '<table style="border-collapse:collapse;">'
        "<tr style='background:#f0f0f0'>"
        "<th style='padding:4px 12px 4px 0;text-align:left'>Agreement ID</th>"
        "<th style='padding:4px 0;text-align:left'>Error</th></tr>"
        f"{rows_errors}</table>"
        "<h3 style='margin-bottom:4px;color:#CC0000;'>Data Quality Issues</h3>"
        '<table style="border-collapse:collapse;margin-bottom:16px;">'
        "<tr style='background:#f0f0f0'>"
        "<th style='padding:4px 8px;text-align:left'>Agreement ID</th>"
        "<th style='padding:4px 8px;text-align:left'>Reseller</th>"
        "<th style='padding:4px 8px;text-align:left'>Data Quality Issues</th>"
        "<th style='padding:4px 8px;text-align:left'>Changed Fields</th>"
        "<th style='padding:4px 8px;text-align:left'>Partial Opt-Out</th>"
        "<th style='padding:4px 8px;text-align:left'>PDF Review</th></tr>"
        f"{rows_dq}</table>"
        '<hr style="margin-top:30px;border:none;border-top:1px solid #ccc;">'
        '<p style="font-size:11px;color:#888;">Automated notification — PSTN Adobe Sign Widget Monitor</p>'
        "</body></html>"
    )

    payload = {
        "message": {
            "subject": f"PSTN Widget Monitor Summary — {run_time}",
            "body": {"contentType": "HTML", "content": html_body},
            "toRecipients": _parse_email_list(PS_SUMMARY_EMAIL),
        }
    }

    try:
        token = _get_graph_token()
        endpoint = f"https://graph.microsoft.com/v1.0/users/{GRAPH_SENDER_MAILBOX}/sendMail"
        headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
        resp = requests.post(endpoint, headers=headers, json=payload, timeout=60)
        if resp.status_code == 202:
            logger.info("✅  PS summary email sent to %s", PS_SUMMARY_EMAIL)
        else:
            logger.error("❌  Failed to send PS summary email: HTTP %s — %s", resp.status_code, resp.text)
    except Exception as exc:
        logger.error("❌  Exception sending PS summary email: %s", exc)


# ── XLS Submissions Report ────────────────────────────────────────────────────
_REPORT_FILE = Path(__file__).parent / "output" / "widget_submissions_report.xlsx"
_REPORT_HEADERS = [
    "Date/Time Received",
    "Agreement ID",
    "Reseller Name",
    "Gamma Account Number",
    "Signer Email",
    "Status",
    "Full Opt-Out",
    "Confirm RWP Remain",
    "Confirm Amendments",
    "Partial Opt-Out",
    "Data Quality Issues",
    "Changed Fields",
    "PDF Human Review",
    "Attachment IDs",
    "Signed PDF",
    "Data Quality Report",
]


def _append_to_xls_report(report_rows: list[dict]) -> None:
    """Append new submission rows to the cumulative XLS report, creating it if needed."""
    if not report_rows:
        return

    if _REPORT_FILE.exists():
        wb = openpyxl.load_workbook(_REPORT_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Submissions"
        ws.append(_REPORT_HEADERS)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        # Freeze header row
        ws.freeze_panes = "A2"

    for r in report_rows:
        has_issues = r.get("data_quality_issues", False)
        is_partial = r.get("partial_opt_out", False)
        is_pdf_review = r.get("pdf_human_review", False)
        row_fill = None
        if has_issues:
            from openpyxl.styles import PatternFill
            row_fill = PatternFill("solid", fgColor="FFCCCC")
        elif is_partial or r.get("changed_fields", ""):
            from openpyxl.styles import PatternFill
            row_fill = PatternFill("solid", fgColor="FFE5B4")

        new_row = ws.append([
            r["received_at"],
            r["agreement_id"],
            r["reseller_name"],
            r["account_number"],
            r["signer_email"],
            r["status"],
            "Yes" if r["opt_out_full"] else "No",
            "Yes" if r["rwp_remain"] else "No",
            "Yes" if r["confirm_amendments"] else "No",
            "Yes" if is_partial else "No",
            "Yes" if has_issues else "No",
            r.get("changed_fields", ""),
            "Yes" if is_pdf_review else "No",
            r["attachment_ids"],
            r["signed_pdf"],
            r.get("data_quality_report", ""),
        ])
        if row_fill:
            for cell in ws[ws.max_row]:
                cell.fill = row_fill

    _REPORT_FILE.parent.mkdir(exist_ok=True)
    wb.save(_REPORT_FILE)
    logger.info(
        "✅  XLS report updated → %s  (%d new row(s))",
        _REPORT_FILE.name, len(report_rows),
    )


def _send_report_email(report_rows: list[dict]) -> None:
    """Email the XLS submissions report as an attachment to CHASER_EMAIL."""
    if not CHASER_EMAIL or not report_rows or not _REPORT_FILE.exists():
        return
    run_time = datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC")
    try:
        token = _get_graph_token()
        data = _REPORT_FILE.read_bytes()
        attachment = {
            "@odata.type": "#microsoft.graph.fileAttachment",
            "name": _REPORT_FILE.name,
            "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "contentBytes": base64.b64encode(data).decode(),
        }
        payload = {
            "message": {
                "subject": f"PSTN Widget Submissions Report — {run_time}",
                "body": {
                    "contentType": "HTML",
                    "content": (
                        '<html><body style="font-family:Arial,sans-serif;font-size:14px;color:#333;">'
                        "<p>Please find attached the latest PSTN Voice Migration widget submissions report.</p>"
                        f"<p><strong>{len(report_rows)}</strong> new submission(s) processed in this run.</p>"
                        '<hr style="margin-top:30px;border:none;border-top:1px solid #ccc;">'
                        '<p style="font-size:11px;color:#888;">Automated notification — PSTN Adobe Sign Widget Monitor</p>'
                        "</body></html>"
                    ),
                },
                "toRecipients": _parse_email_list(CHASER_EMAIL),
                "attachments": [attachment],
            }
        }
        endpoint = f"https://graph.microsoft.com/v1.0/users/{GRAPH_SENDER_MAILBOX}/sendMail"
        headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
        resp = requests.post(endpoint, headers=headers, json=payload, timeout=60)
        if resp.status_code == 202:
            logger.info("✅  Report email sent to %s", CHASER_EMAIL)
        else:
            logger.error("❌  Failed to send report email: HTTP %s — %s", resp.status_code, resp.text)
    except Exception as exc:
        logger.error("❌  Exception sending report email: %s", exc)


# ── Main ──────────────────────────────────────────────────────────────────────

def run_widget_monitor() -> dict:
    """
    Poll the Adobe Sign widget for new completed submissions and process each one.

    Returns:
        Summary dict: {processed, skipped, errors}
    """
    logger.info("=" * 60)
    logger.info("Widget monitor started")
    logger.info("=" * 60)

    if not ADOBE_SIGN_WIDGET_ID or ADOBE_SIGN_WIDGET_ID.startswith("YOUR_"):
        logger.error(
            "ADOBE_SIGN_WIDGET_ID is not configured. "
            "Set it in .env to the widget ID from your webform URL (wid=...)."
        )
        return {"processed": 0, "skipped": 0, "errors": 1}

    # GitHub pre-run backup of circuit master (if configured and update enabled)
    github_backup_sha: Optional[str] = None
    github_backup_failed = False
    if UPDATE_MASTER_DATA and GITHUB_REPO and GITHUB_TOKEN and CIRCUIT_MASTER_PATH:
        from pathlib import Path as _Path
        github_backup_sha = backup_master_to_github(_Path(CIRCUIT_MASTER_PATH))
        if github_backup_sha:
            logger.info("GitHub master backup committed: %s", github_backup_sha)
        else:
            github_backup_failed = True
            logger.warning(
                "GitHub backup of circuit master failed or was skipped. "
                "Master updates will still proceed — restore manually from last known backup if needed."
            )

    client = AdobeSignClient()
    widget_state = _load_widget_state()

    # Fetch all agreements submitted via the widget
    try:
        agreements = client.list_widget_agreements(ADOBE_SIGN_WIDGET_ID)
    except AdobeSignError as exc:
        logger.error("Failed to fetch widget agreements: %s", exc)
        return {"processed": 0, "skipped": 0, "errors": 1}

    logger.info("Total widget submissions found: %d", len(agreements))

    processed_records: list[dict] = []
    error_records: list[dict] = []
    report_rows: list[dict] = []
    skipped = 0

    for ag in agreements:
        agreement_id = ag.get("agreementId") or ag.get("id", "")
        status = ag.get("status", "")

        if not agreement_id:
            continue

        # Skip already processed
        if agreement_id in widget_state:
            skipped += 1
            continue

        # Skip non-completed submissions (still in progress / awaiting signature)
        if status not in _COMPLETED_STATUSES:
            logger.debug("Skipping agreement %s — status: %s", agreement_id, status)
            skipped += 1
            continue

        logger.info("Processing agreement %s (status: %s)", agreement_id, status)

        # Get full agreement details to extract signer name/email
        try:
            agreement_detail = client.get_agreement(agreement_id)
        except AdobeSignError as exc:
            logger.error("Could not fetch agreement details for %s: %s", agreement_id, exc)
            error_records.append({"agreement_id": agreement_id, "error": str(exc)})
            continue

        signer_name = _get_signer_name(agreement_detail)
        signer_email = _get_signer_email(agreement_detail)

        # Process the agreement (downloads form data, attachments, signed PDF)
        try:
            result = process_signed_agreement(agreement_id, signer_name)
        except Exception as exc:
            logger.error("Processing failed for %s (%s): %s", signer_name, agreement_id, exc)
            error_records.append({"agreement_id": agreement_id, "error": str(exc)})
            continue

        # Build the audit record from processing result
        audit_data = {
            "signer_name": result.get("reseller_name") or signer_name,
            "signer_email": signer_email,
            "account_ref": result.get("account_number", ""),
            "migration_status": status,
            "opt_out_full": result.get("opt_out_full", False),
            "rwp_remain": result.get("rwp_remain", False),
            "confirm_amendments": result.get("confirm_amendments", False),
            "processed_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
        }

        # Write to Audit tab in master Excel
        try_write_to_audit_tab(audit_data, agreement_id)

        # Save to widget state
        widget_state[agreement_id] = {
            "signer_name": signer_name,
            "signer_email": signer_email,
            "status": status,
            "opt_out_full": result.get("opt_out_full", False),
            "processed_at": datetime.now(timezone.utc).isoformat(),
            "errors": result.get("errors", []),
        }
        _save_widget_state(widget_state)

        # Summarise data quality validation results (used by both report_rows and processed_records)
        val_results = result.get("validation_results", [])
        changed_fields_summary = "; ".join(
            f"{r.circuit_id}:{col}"
            for v in val_results
            for r in v.row_results
            for col in r.mutable_changes
        )
        pdf_human_review = any(v.pdf_human_review_required for v in val_results)

        processed_records.append({
            "agreement_id": agreement_id,
            "signer_name": signer_name,
            "signer_email": signer_email,
            "reseller_name": result.get("reseller_name", "") or signer_name,
            "opt_out_full": result.get("opt_out_full", False),
            "partial_opt_out": result.get("is_partial_opt_out", False),
            "data_quality_issues": result.get("has_data_quality_issues", False),
            "changed_fields": changed_fields_summary,
            "pdf_human_review": pdf_human_review,
        })

        # Collect row for XLS report
        received_at = datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M:%S UTC")

        report_rows.append({
            "received_at": received_at,
            "agreement_id": agreement_id,
            "reseller_name": result.get("reseller_name", "") or signer_name,
            "account_number": result.get("account_number", ""),
            "signer_email": signer_email,
            "status": status,
            "opt_out_full": result.get("opt_out_full", False),
            "rwp_remain": result.get("rwp_remain", False),
            "confirm_amendments": result.get("confirm_amendments", False),
            "partial_opt_out": result.get("is_partial_opt_out", False),
            "data_quality_issues": result.get("has_data_quality_issues", False),
            "changed_fields": changed_fields_summary,
            "pdf_human_review": pdf_human_review,
            "attachment_ids": ", ".join(result.get("attachment_ids", [])),
            "signed_pdf": Path(result["signed_pdf_file"]).name if result.get("signed_pdf_file") else "",
            "data_quality_report": Path(result["data_quality_report_file"]).name if result.get("data_quality_report_file") else "",
        })
        logger.info("✅  Processed: %s <%s>", signer_name, signer_email)

    summary = {
        "processed": len(processed_records),
        "skipped": skipped,
        "errors": len(error_records),
    }
    logger.info(
        "Widget monitor complete — processed: %d | skipped: %d | errors: %d",
        summary["processed"], summary["skipped"], summary["errors"],
    )

    # Flush all in-memory master cell writes to disk in a single wb.save() call.
    # This is deferred from _update_master to avoid saving once per agreement
    # on a 19k-row file (30-60s each).
    if UPDATE_MASTER_DATA:
        try:
            flush_master_writes()
        except Exception as exc:
            logger.error("flush_master_writes raised unexpectedly: %s", exc, exc_info=True)

    # Write XLS submissions report
    _append_to_xls_report(report_rows)

    # Email report to chaser mailbox
    _send_report_email(report_rows)

    # Send PS summary email
    _send_ps_summary(processed_records, error_records, skipped)

    return summary


def _reset_report() -> None:
    """Delete the cumulative submissions report and clear widget state, ready for a fresh go-live run."""
    import shutil
    reset_count = 0
    if _REPORT_FILE.exists():
        _REPORT_FILE.unlink()
        logger.info("🗑  Submissions report deleted: %s", _REPORT_FILE)
        reset_count += 1
    else:
        logger.info("Submissions report not found — nothing to delete.")

    state_file = Path(__file__).parent / "input" / "widget_state.json"
    if state_file.exists():
        state_file.write_text("{}", encoding="utf-8")
        logger.info("🗑  Widget state reset to {}: %s", state_file)
        reset_count += 1
    else:
        logger.info("Widget state file not found — nothing to reset.")

    if reset_count:
        logger.info("✅  Reset complete — %d file(s) cleared. Ready for go-live.", reset_count)
    else:
        logger.info("Nothing to reset.")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="PSTN Adobe Sign Widget Monitor")
    parser.add_argument(
        "--reset-report",
        action="store_true",
        help="Delete the submissions report and clear widget state (use before go-live to remove test records).",
    )
    args = parser.parse_args()
    if args.reset_report:
        _reset_report()
    else:
        run_widget_monitor()
